import genanki
from openpyxl import load_workbook
import random
import tkinter as tk
from tkinter import filedialog, messagebox


def create_anki_deck(excel_file_path, deck_name, output_file_name):
    deck_id = random.randint(1, 2 ** 30)

    radio_model = genanki.Model(
        542046268,
        'radio',
        fields=[
            {'name': 'Question'},
            {'name': 'Answer'},
            {'name': 'analysis'},
            {'name': 'score'},
            {'name': 'type'},
            {'name': 'questionNum'},
            {'name': 'options'},
        ],
        templates=[
            {
                'name': 'Card 1',
                'qfmt': '''
    <script>
        // v1.1.8 - https://github.com/SimonLammer/anki-persistence/blob/584396fea9dea0921011671a47a0fdda19265e62/script.js
        if (void 0 === window.Persistence) { var e = "github.com/SimonLammer/anki-persistence/", t = "_default"; if (window.Persistence_sessionStorage = function () { var i = !1; try { "object" == typeof window.sessionStorage && (i = !0, this.clear = function () { for (var t = 0; t < sessionStorage.length; t++) { var i = sessionStorage.key(t); 0 == i.indexOf(e) && (sessionStorage.removeItem(i), t--) } }, this.setItem = function (i, n) { void 0 == n && (n = i, i = t), sessionStorage.setItem(e + i, JSON.stringify(n)) }, this.getItem = function (i) { return void 0 == i && (i = t), JSON.parse(sessionStorage.getItem(e + i)) }, this.removeItem = function (i) { void 0 == i && (i = t), sessionStorage.removeItem(e + i) }, this.getAllKeys = function () { for (var t = [], i = Object.keys(sessionStorage), n = 0; n < i.length; n++) { var s = i[n]; 0 == s.indexOf(e) && t.push(s.substring(e.length, s.length)) } return t.sort() }) } catch (n) { } this.isAvailable = function () { return i } }, window.Persistence_windowKey = function (i) { var n = window[i], s = !1; "object" == typeof n && (s = !0, this.clear = function () { n[e] = {} }, this.setItem = function (i, s) { void 0 == s && (s = i, i = t), n[e][i] = s }, this.getItem = function (i) { return void 0 == i && (i = t), void 0 == n[e][i] ? null : n[e][i] }, this.removeItem = function (i) { void 0 == i && (i = t), delete n[e][i] }, this.getAllKeys = function () { return Object.keys(n[e]) }, void 0 == n[e] && this.clear()), this.isAvailable = function () { return s } }, window.Persistence = new Persistence_sessionStorage, Persistence.isAvailable() || (window.Persistence = new Persistence_windowKey("py")), !Persistence.isAvailable()) { var i = window.location.toString().indexOf("title"), n = window.location.toString().indexOf("main", i); i > 0 && n > 0 && n - i < 10 && (window.Persistence = new Persistence_windowKey("qt")) } }
    </script>
    <div class="question">
        <div class="tag">{{type}}</div>
        <p class="header">
            <p class="question-number">题号：{{questionNum}}  分值：{{score}}</p>
            <p class="question-content">{{Question}}</p>
        </p>
        <div class="options" id="optionsWrap">
            {{options}}
        </div>
    </div>
    <script>
        function flipToBack() {
            if (typeof pycmd !== "undefined") {
                pycmd("ans")
            } else if (typeof study !== "undefined") {
                study.drawAnswer()
            } else if (typeof AnkiDroidJS !== "undefined") {
                showAnswer()
            } else if (window.anki && window.sendMessage2) {
                window.sendMessage2("ankitap", "midCenter")
            }
        }
        function checkThis(e) {
            flipToBack()
            var isCompete = document.getElementById('optionsWrap').getAttribute('isCompete');
            if (isCompete == 'true') {
                checkLabel.style.backgroundColor = 'red';
                return;
            }
            document.getElementById('optionsWrap').setAttribute('isCompete', 'true');
            var Answer = "{{Answer}}";
            // 获取元素的属性
            var checkAnswer = e.getAttribute('answer');
            if (Persistence.isAvailable()) {
                Persistence.setItem('checkAnswer', checkAnswer)
            }
            window.checkAnswer = checkAnswer
            // /设置label的样式
            var checkLabel = document.getElementById(checkAnswer)
            var answerLabel = document.getElementById(Answer)
            document.getElementById('test').innerHtml = '选择过了'
            if (checkAnswer == Answer) {
                console.log('正确');
                checkLabel.style.backgroundColor = 'rgb(225, 243, 216)';
                checkLabel.style.borderColor = 'rgb(225, 243, 216)';
            } else {
                console.log('错误');
                answerLabel.style.backgroundColor = 'rgb(225, 243, 216)';
                answerLabel.style.borderColor = 'rgb(225, 243, 216)';
                checkLabel.style.backgroundColor = 'rgb(253, 226, 226)';
                checkLabel.style.borderColor = 'rgb(253, 226, 226)';
            }

        }
    </script>
 ''',
                'afmt': '''
                   <script>
// v1.1.8 - https://github.com/SimonLammer/anki-persistence/blob/584396fea9dea0921011671a47a0fdda19265e62/script.js
if(void 0===window.Persistence){var e="github.com/SimonLammer/anki-persistence/",t="_default";if(window.Persistence_sessionStorage=function(){var i=!1;try{"object"==typeof window.sessionStorage&&(i=!0,this.clear=function(){for(var t=0;t<sessionStorage.length;t++){var i=sessionStorage.key(t);0==i.indexOf(e)&&(sessionStorage.removeItem(i),t--)}},this.setItem=function(i,n){void 0==n&&(n=i,i=t),sessionStorage.setItem(e+i,JSON.stringify(n))},this.getItem=function(i){return void 0==i&&(i=t),JSON.parse(sessionStorage.getItem(e+i))},this.removeItem=function(i){void 0==i&&(i=t),sessionStorage.removeItem(e+i)},this.getAllKeys=function(){for(var t=[],i=Object.keys(sessionStorage),n=0;n<i.length;n++){var s=i[n];0==s.indexOf(e)&&t.push(s.substring(e.length,s.length))}return t.sort()})}catch(n){}this.isAvailable=function(){return i}},window.Persistence_windowKey=function(i){var n=window[i],s=!1;"object"==typeof n&&(s=!0,this.clear=function(){n[e]={}},this.setItem=function(i,s){void 0==s&&(s=i,i=t),n[e][i]=s},this.getItem=function(i){return void 0==i&&(i=t),void 0==n[e][i]?null:n[e][i]},this.removeItem=function(i){void 0==i&&(i=t),delete n[e][i]},this.getAllKeys=function(){return Object.keys(n[e])},void 0==n[e]&&this.clear()),this.isAvailable=function(){return s}},window.Persistence=new Persistence_sessionStorage,Persistence.isAvailable()||(window.Persistence=new Persistence_windowKey("py")),!Persistence.isAvailable()){var i=window.location.toString().indexOf("title"),n=window.location.toString().indexOf("main",i);i>0&&n>0&&n-i<10&&(window.Persistence=new Persistence_windowKey("qt"))}}
</script>
                        {{FrontSide}}
                         <div class="answer">正确答案： {{Answer}}</div>
                         <p> 解析： {{analysis}} </p>

                         <script>
        var checkAnswer= Persistence.getItem('checkAnswer') //选择的答案
        var Answer = "{{Answer}}" //正确答案
        var checkLabel = document.getElementById(checkAnswer)
        var answerLabel = document.getElementById(Answer)
        if (checkAnswer == Answer) {
            checkLabel.style.backgroundColor = 'rgb(225, 243, 216)';
            checkLabel.style.borderColor = 'rgb(225, 243, 216)';
        } else {
            answerLabel.style.backgroundColor = 'rgb(225, 243, 216)';
            answerLabel.style.borderColor = 'rgb(225, 243, 216)';
            checkLabel.style.backgroundColor = 'rgb(253, 226, 226)';
            checkLabel.style.borderColor = 'rgb(253, 226, 226)';
        }
        function checkThis(e) {}
        Persistence.clear();  


    </script>
                    ''',
            },
        ],
        css='''

        .tag {
            display: inline-block;
            font-size: 12px;
            padding: 2px 5px;
            color: #1772F6;
            background-color: rgb(217, 236, 255);
            /*color: #fde2e2; 红色 
            color: #e1f3d8;  蓝色 */

            border-radius: 5px;
        }

        .question {
            border: 1px solid #ccc;
            margin: 20px auto;
            padding: 10px;
            width: 100%;
            box-sizing: border-box;

        }

        .options {
            margin-bottom: 15px;
            display: flex;
            flex-direction: column;
        }

        .option {
            margin-bottom: 10px;
            box-sizing: border-box;
            padding: 10px;
            display: flex;
            overflow-wrap: break-word;
            cursor: pointer;
            height: 100%;
            border: 1px solid #ccc;
            font-size: 14px;
            color: #333;
            border-radius: 5px;
        }



        .option p {
            margin: 0;
            margin-right: 10px;
        }
        '''
    )

    check_model = genanki.Model(
        504743187,
        'check',
        fields=[
            {'name': 'Question'},
            {'name': 'Answer'},
            {'name': 'analysis'},
            {'name': 'score'},
            {'name': 'type'},
            {'name': 'questionNum'},
            {'name': 'options'},
        ],
        templates=[
            {
                'name': 'Card 1',
                'qfmt': '''
                        <script>
            // v1.1.8 - https://github.com/SimonLammer/anki-persistence/blob/584396fea9dea0921011671a47a0fdda19265e62/script.js
            if (void 0 === window.Persistence) { var e = "github.com/SimonLammer/anki-persistence/", t = "_default"; if (window.Persistence_sessionStorage = function () { var i = !1; try { "object" == typeof window.sessionStorage && (i = !0, this.clear = function () { for (var t = 0; t < sessionStorage.length; t++) { var i = sessionStorage.key(t); 0 == i.indexOf(e) && (sessionStorage.removeItem(i), t--) } }, this.setItem = function (i, n) { void 0 == n && (n = i, i = t), sessionStorage.setItem(e + i, JSON.stringify(n)) }, this.getItem = function (i) { return void 0 == i && (i = t), JSON.parse(sessionStorage.getItem(e + i)) }, this.removeItem = function (i) { void 0 == i && (i = t), sessionStorage.removeItem(e + i) }, this.getAllKeys = function () { for (var t = [], i = Object.keys(sessionStorage), n = 0; n < i.length; n++) { var s = i[n]; 0 == s.indexOf(e) && t.push(s.substring(e.length, s.length)) } return t.sort() }) } catch (n) { } this.isAvailable = function () { return i } }, window.Persistence_windowKey = function (i) { var n = window[i], s = !1; "object" == typeof n && (s = !0, this.clear = function () { n[e] = {} }, this.setItem = function (i, s) { void 0 == s && (s = i, i = t), n[e][i] = s }, this.getItem = function (i) { return void 0 == i && (i = t), void 0 == n[e][i] ? null : n[e][i] }, this.removeItem = function (i) { void 0 == i && (i = t), delete n[e][i] }, this.getAllKeys = function () { return Object.keys(n[e]) }, void 0 == n[e] && this.clear()), this.isAvailable = function () { return s } }, window.Persistence = new Persistence_sessionStorage, Persistence.isAvailable() || (window.Persistence = new Persistence_windowKey("py")), !Persistence.isAvailable()) { var i = window.location.toString().indexOf("title"), n = window.location.toString().indexOf("main", i); i > 0 && n > 0 && n - i < 10 && (window.Persistence = new Persistence_windowKey("qt")) } }
        </script>
        <div class="question">
            <div class="tag">{{type}}</div>
            <p class="header">
                <p class="question-number">题号：{{questionNum}}  分值：{{score}}</p>
                <p class="question-content">{{Question}}</p>
            </p>
            <div class="options" id="optionsWrap">
                {{options}}
            </div>
        </div>


        <script>

            function flipToBack() {
                if (typeof pycmd !== "undefined") {
                    pycmd("ans")
                } else if (typeof study !== "undefined") {
                    study.drawAnswer()
                } else if (typeof AnkiDroidJS !== "undefined") {
                    showAnswer()
                } else if (window.anki && window.sendMessage2) {
                    window.sendMessage2("ankitap", "midCenter")
                }
            }

            var checkAnswerList = []
            Answer = Answer.split(',');
            // 获取元素的属性
             function checkThis(e) {
            var checkAnswer = e.getAttribute('answer');
            var checkLabel = document.getElementById(checkAnswer)
            if(checkAnswerList.indexOf(checkAnswer) != -1){
                checkLabel.style.backgroundColor = '#fde2e2';
                checkLabel.style.borderColor = '#fde2e2';
                checkAnswerList.splice(checkAnswerList.indexOf(checkAnswer),1)
            }else{
                checkLabel.style.backgroundColor = '#e1f3d8';
                checkLabel.style.borderColor = '#e1f3d8';
                checkAnswerList.push(checkAnswer)
            }

            if (Persistence.isAvailable()) {
                Persistence.setItem('checkAnswerList', checkAnswerList)
            }
            if (Answer.length != checkAnswerList.length) {
                return;
            } else {
                flipToBack()
            }
        }



        </script>
                            ''',
                'afmt': '''

        <script>
            // v1.1.8 - https://github.com/SimonLammer/anki-persistence/blob/584396fea9dea0921011671a47a0fdda19265e62/script.js
            if (void 0 === window.Persistence) { var e = "github.com/SimonLammer/anki-persistence/", t = "_default"; if (window.Persistence_sessionStorage = function () { var i = !1; try { "object" == typeof window.sessionStorage && (i = !0, this.clear = function () { for (var t = 0; t < sessionStorage.length; t++) { var i = sessionStorage.key(t); 0 == i.indexOf(e) && (sessionStorage.removeItem(i), t--) } }, this.setItem = function (i, n) { void 0 == n && (n = i, i = t), sessionStorage.setItem(e + i, JSON.stringify(n)) }, this.getItem = function (i) { return void 0 == i && (i = t), JSON.parse(sessionStorage.getItem(e + i)) }, this.removeItem = function (i) { void 0 == i && (i = t), sessionStorage.removeItem(e + i) }, this.getAllKeys = function () { for (var t = [], i = Object.keys(sessionStorage), n = 0; n < i.length; n++) { var s = i[n]; 0 == s.indexOf(e) && t.push(s.substring(e.length, s.length)) } return t.sort() }) } catch (n) { } this.isAvailable = function () { return i } }, window.Persistence_windowKey = function (i) { var n = window[i], s = !1; "object" == typeof n && (s = !0, this.clear = function () { n[e] = {} }, this.setItem = function (i, s) { void 0 == s && (s = i, i = t), n[e][i] = s }, this.getItem = function (i) { return void 0 == i && (i = t), void 0 == n[e][i] ? null : n[e][i] }, this.removeItem = function (i) { void 0 == i && (i = t), delete n[e][i] }, this.getAllKeys = function () { return Object.keys(n[e]) }, void 0 == n[e] && this.clear()), this.isAvailable = function () { return s } }, window.Persistence = new Persistence_sessionStorage, Persistence.isAvailable() || (window.Persistence = new Persistence_windowKey("py")), !Persistence.isAvailable()) { var i = window.location.toString().indexOf("title"), n = window.location.toString().indexOf("main", i); i > 0 && n > 0 && n - i < 10 && (window.Persistence = new Persistence_windowKey("qt")) } }
        </script>

        {{FrontSide}}
        <hr id="answer">

        <div class="answer">正确答案： {{Answer}}</div>
        <p> 解析： {{analysis}} </p>

        <script>
            var checkAnswer = Persistence.getItem('checkAnswerList') //选择的答案
            var Answer = "{{Answer}}" //正确答案
            Answer = Answer.split(',');
            for (var i = 0; i < checkAnswer.length; i++) {
                var checkLabel = document.getElementById(checkAnswer[i])
                checkLabel.style.backgroundColor = '#fde2e2';
                checkLabel.style.borderColor = '#fde2e2';
            }
            for (var j = 0; j < Answer.length; j++) {
                var answerLabel = document.getElementById(Answer[j])
                answerLabel.style.backgroundColor = '#e1f3d8';
                answerLabel.style.borderColor = '#e1f3d8';
            }
            function checkThis(e) { }
            Persistence.clear();

        </script>
                        ''',
            },
        ],
        css='''  

        .tag {
            display: inline-block;
            font-size: 12px;
            padding: 2px 5px;
            color: #1772F6;
            background-color: rgb(217, 236, 255);
            /*color: #fde2e2; 红色 
            color: #e1f3d8;  蓝色 */

            border-radius: 5px;
        }

        .question {
            border: 1px solid #ccc;
            margin: 20px auto;
            padding: 10px;
            width: 100%;
            box-sizing: border-box;

        }

        .options {
            margin-bottom: 15px;
            display: flex;
            flex-direction: column;
        }

        .option {
            margin-bottom: 10px;
            box-sizing: border-box;
            padding: 10px;
            display: flex;
            overflow-wrap: break-word;
            cursor: pointer;
            height: 100%;
            border: 1px solid #ccc;
            font-size: 14px;
            color: #333;
            border-radius: 5px;
        }



        .option p {
            margin: 0;
            margin-right: 10px;
        }
                '''
    )

    my_deck = genanki.Deck(
        deck_id,
        deck_name
    )

    try:
        workbook = load_workbook(excel_file_path)

        for sheet in workbook.worksheets:
            total_rows = sheet.max_row
            for index in range(2, total_rows + 1):
                row = sheet[index]
                if row[2].value:
                    questionNum = row[0].value or '没有设置题号'
                    type = row[1].value or '未知题型'
                    question = row[2].value or ''
                    answer = row[3].value or '未设置答案'
                    analysis = row[4].value or '无解析'
                    score = row[5].value or '1'

                    # 获取选项并将其格式化
                    options = [
                        ("A", row[6].value),
                        ("B", row[7].value),
                        ("C", row[8].value),
                        ("D", row[9].value),
                        ("E", row[10].value),
                        ("F", row[11].value),
                    ]

                    # 生成HTML字符串，仅包含有值的选项
                    innerhtml = ""
                    for id, value in options:
                        if value:  # 仅在有值的情况下生成HTML
                            innerhtml += f'''
                            <div class="option" id="{id}" answer="{id}" onclick="checkThis(this)">
                                <p>{id}.</p> <span>{value}</span>
                            </div>
                            '''

                    # 创建 Anki 卡片
                    if '多选' in type:
                        my_note = genanki.Note(
                            model=check_model,
                            fields=[
                                str(question),
                                str(answer),
                                str(analysis),
                                str(score),
                                str(type),
                                str(questionNum),
                                str(innerhtml),
                            ]
                        )
                    else:
                        my_note = genanki.Note(
                            model=radio_model,
                            fields=[
                                str(question),
                                str(answer),
                                str(analysis),
                                str(score),
                                str(type),
                                str(questionNum),
                                str(innerhtml),
                            ]
                        )
                    my_deck.add_note(my_note)
        genanki.Package(my_deck).write_to_file(output_file_name)
        messagebox.showinfo('成功', f'Anki包已生成：{output_file_name}')

    except Exception as e:
        messagebox.showerror('错误', f'处理文件时发生错误: {e}')
        print(e)


def select_excel_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if file_path:
        excel_file_entry.delete(0, tk.END)
        excel_file_entry.insert(0, file_path)

def generate_deck():
    excel_file_path = excel_file_entry.get()
    deck_name = deck_name_entry.get()
    output_file_name = f'{deck_name}.apkg'

    if not excel_file_path or not deck_name:
        messagebox.showwarning('警告', '请填写所有必填项。')
        return

    create_anki_deck(excel_file_path, deck_name, output_file_name)

# 创建主窗口
root = tk.Tk()
root.title('Anki 牌组生成器')

# Excel文件选择
tk.Label(root, text='选择Excel文件:').grid(row=0, column=0, padx=10, pady=10)
excel_file_entry = tk.Entry(root, width=40)
excel_file_entry.grid(row=0, column=1, padx=10, pady=10)
tk.Button(root, text='浏览', command=select_excel_file).grid(row=0, column=2, padx=10, pady=10)

# 牌组名称输入
tk.Label(root, text='输入牌组名称:').grid(row=1, column=0, padx=10, pady=10)
deck_name_entry = tk.Entry(root, width=40)
deck_name_entry.grid(row=1, column=1, padx=10, pady=10)

# 生成按钮
tk.Button(root, text='生成Anki牌组', command=generate_deck).grid(row=2, column=1, padx=10, pady=10)

# 运行主循环
root.mainloop()
